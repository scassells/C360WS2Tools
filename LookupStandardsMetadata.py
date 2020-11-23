# LookupStandardsMetadata
#
# The MIT License (MIT)
#  
#   Permission is hereby granted, free of charge, to any person obtaining a copy of this software and 
#   associated documentation files (the "Software"), to deal in the Software without restriction, 
#   including without limitation the rights to use, copy, modify, merge, publish, distribute, 
#   sublicense, and/or sell copies of the Software, and to permit persons to whom the Software is 
#   furnished to do so, subject to the following conditions:
#   The above copyright notice and this permission notice shall be included in all copies or 
#   substantial portions of the Software.
# 
#   THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR IMPLIED, INCLUDING BUT 
#   NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE AND 
#   NONINFRINGEMENT. IN NO EVENT SHALL THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, 
#   DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM, OUT 
#   OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE SOFTWARE.

import json, os, fnmatch, requests, getpass, logging
import config
from typing import Dict, Any, Union

from openpyxl import load_workbook

# looks up the domain in CLIB then uses the cmap to make updates to the variable definitions
cdisc_prod_LibraryURL = "https://library.cdisc.org/api"
hdr4xml = {'accept': 'application/xml','api-key': ''}

testhdr = {'api-key': config.api_key}

def lookupStandardSDTM(standard, version, domain, tokenList) ->dict:
	"""

	:rtype:
	"""
	uname = tokenList[0]
	passwd = tokenList[1]
	baseURL = tokenList[2]
	stdVersURL =  baseURL + '/mdr/' + standard + '/' + version + '/datasets/'
	dsURL = lambda dsURL, dsName: dsURL + dsName
	varURL = lambda dsURL, varName : dsURL + '/variables/' + varName

	domainURL = dsURL(stdVersURL, domain)
	domainCLIB = testRequest(domainURL, uname, passwd)
	dmdict = {}
	dmdict["OID"] = "C360.SDTM." + standard + '_' + version + "." + domain
	dmdict["Name"] = domain
	dmdict["Label"] = domainCLIB["label"]
	dmdict["Comment"] = domainCLIB["description"]
	dictLinks = domainCLIB["_links"]
	classLink = dictLinks["parentClass"]
	dclass = classLink["title"].split()[0]
	dmdict["Class"] = dclass
	dmdict["datasetStructure"] = domainCLIB["datasetStructure"]
	# we will use the set of variables specified in the cmap in the bc
	clibVarList = domainCLIB["datasetVariables"]

	c360varList = []
	clRefs = []
	c360clList = []
	clibOrdinal = 1
	for v in clibVarList:
		cvarIndex = indexClibVar(v["name"], clibVarList)
		logging.info("Look up of %s returned %s.", v["name"], cvarIndex)
		if cvarIndex >= 0:
			c360varDef = getSDTMClibVar(v["name"], clibVarList)
			c360clDef = clibCodeList(clibVarList[cvarIndex], tokenList)
			if "href" in c360clDef:
				clhref = c360clDef["href"]
				if clhref not in clRefs:   # some are repeated only copy once
					clRefs.append(clhref)
					c360clList.append(c360clDef)
					logging.info(" Add codelist for " + v["name"])
			else:
				logging.debug("no href" + c360clDef["name"])
			logging.info("Looked up " +  v["name"])
			c360varList.append(c360varDef)
			clibOrdinal = clibOrdinal + 1
		else:
			logging.info("Did not find Clib info for " + v["name"] + " in " + domainURL)
		# to do process codelist, method, vlm, codelist subsets
	
	if len(c360varList) > 0:
		outcome = "For "  + standard + "," + version + domain +  " has " + str(len(c360varList)) + " decs"
		logging.info(outcome)
		dmdict["bcVarsets"] = c360varList
	
	defInfo = {}
	defInfo["ProdVersion"] = standard + "-" + version
	defInfo["Standard"] = standard
	defInfo["StdVersion"] = version
	defInfo["title"] = domain + " Biomedical Concepts CDISC 360"
	defInfo["domain"] = dmdict
	defInfo["codelists"] = c360clList
	fname = domain + "_SDTMClib.json"
	logging.debug("Output file " + fname)
	with open(fname, "w", encoding="utf-8") as fo:
		fo.write(json.dumps(defInfo))
	return(defInfo)
	
	
def lookupCDASHIG(version, cmap_report, tokenList, bcList) ->dict:
	uname = tokenList[0]
	passwd = tokenList[1]
	baseURL = tokenList[2]
	stdVersURL =  baseURL + '/mdr/cdashig/' + version + '/domains/'
	domainURL = lambda dURL, dName: dURL + dName
	fieldURL = lambda dURL, fldName : dURL + '/fields/' + fldName
	# this processes only 1 cmap so it creates the ODM content for only 1 dataset
	# creates a list of domains because that what the existing writeXML expects
	wb = load_workbook(cmap_report)
	sheet = wb["CMAP Nodes"]
	if len(bcList) == 0:
	   doConcepts = False
	else:
	   doConcepts = True
	
	domain = getSDTMDomain(sheet) 
	bcname = getBC(sheet) + "-CDASH"
	dsURL = domainURL(stdVersURL, domain)
	logging.info(dsURL)
	domainCLIB = testRequest(dsURL, uname, passwd)
	dmdict = {}
	dmdict["OID"] = "C360.CDASH.DOMAIN." + domain
	dmdict["Name"] = domain
	dmdict["Label"] = domainCLIB["label"]
	dictLinks = domainCLIB["_links"]
	# we will use the set of variables specified in the cmap in the bc
	dmdict["Class"] = dictLinks["parentClass"] 
	clibFldList = domainCLIB["fields"]
	nfldsDomain = len(clibFldList)
	fldnames = getVarnames(sheet)
	nfldsBC = len(fldnames)
	bcinfo = "Standard " + domain + " has " + str(nfldsDomain) + " BC " + bcname + " has " + str(nfldsBC)
	logging.info(bcinfo)
	codelists = getCodelists(sheet)
	clSubsets = getCLSubsets(sheet)
	clNumbers = "Number of Codelists " + str(len(codelists)) + " up to " + str(len(clSubsets)) + " subsets."
	logging(clNumbers)
	c360fldList = []
	clibOrdinal = 1
	for f in fldnames:
		cfldIndex = indexClibVar(f, clibFldList)
		if cfldIndex >= 0:
			c360fldDef = clibFldList[cfldIndex]
			logging.info(c360fldDef["ordinal"] + " " + c360fldDef["name"])
			c360fldDef["ordinal"] = str(clibOrdinal)
			c360fldList.append(c360fldDef)
			clibOrdinal = clibOrdinal + 1
		# to do process codelist, method, vlm, codelist subsets
	
	outcome = "For " + bcname + "(cdashig," + version + ")" + domain +  " has " + str(len(c360fldList)) + " decs"
	logging.info(outcome)
	dmdict["bcFields"] = c360fldList
	domainList = []
	domainList.append(dmdict)
	
	defInfo = {}
	defInfo["ProdVersion"] = "cdashig-" + version + "_" + bcname
	defInfo["Standard"] = "cdashig"
	defInfo["StdVersion"] = version
	defInfo["title"] = domain + "Biomedical Concepts CDASH IG v" + version
	defInfo["domainList"] = domainList
	defInfo["codelists"] = codelists
	defInfo["clSubsets"] = clSubsets
	fname = domain + "_CDASHClib.json"
	logging.info(fname)
	with open(fname, "w", encoding="utf-8") as fo:
		fo.write(json.dumps(defInfo))
	return(defInfo)

def indexClibVar(vname:str, dsVarList:list) ->  int:
	for i in range(0, len(dsVarList)):
		varDef = dsVarList[i]
		if varDef["name"] == vname:
			return i

	return -1

def getSDTMClibVar(vname:str, dsVarList:list)-> dict:
	clibSDTMdef = {}
	for i in range(0, len(dsVarList)):
		varDef = dsVarList[i]
		if varDef["name"] == vname:
			return varDef
			"""
			linksList = varDef["_links"]
			if "self" in linksList:
				clibSDTMdef[vname] = linksList["self"]
				return clibSDTMdef
				"""

	logging.error("Variable %s not found.", vname)
	return clibSDTMdef

def getADaMClibVar(vname:str, varsets:list) ->dict:
	clibVarDef = {}
	for varset in varsets:
		vsVarList = varset["analysisVariables"]
		vsIdx = indexClibVar(vname, vsVarList)
		if vsIdx >= 0:
			clibVarDef[vname] = vsVarList[vsIdx]
			clibVarDef["varsetName"] = varset["name"]
			return clibVarDef
	return clibVarDef

	
def getSDTMDomain(sheet) -> str:
	nfound = 0
	for r in sheet.iter_rows(min_col=1, max_col=2, values_only=True):
		ctype = r[1]
		if ctype == "SDTM Domain":
		   dname = r[0]
		   return dname
	return

def getADaMdataset(sheet) -> str:
	nfound = 0
	dname = ""
	for r in sheet.iter_rows(min_col=1, max_col=2, values_only=True):
		ctype = r[1]
		if ctype == "Dataset":
		   dname = r[0]
		   return dname
	return dname
	
def getBC(sheet) -> str:
	bcname = ""
	for r in sheet.iter_rows(min_col=1, max_col=2, values_only=True):
		ctype = r[1]
		if ctype == "Observation Concept":
		   bcname = r[0]
	return bcname
	
def getVarnames(sheet) -> list:
	varList = []
	for r in sheet.iter_rows(min_col=1, max_col=2, values_only=True):
		ctype = r[1]
		if ctype == "Data Element Concept":
			varName = r[0]
			varList.append(varName)
	return varList

def getCodelists(sheet) -> list:
	# Some Cmaps, for example DM,  do not include codelists
	codelistList = []
	nfound = 0
	for r in sheet.iter_rows(min_col=1, max_col=2, values_only=True):
		ctype = r[1]
		if ctype == "Codelist":
			codelistName = r[0]
			# print("CodelistName " + codelistName)
			if codelistName not in codelistList or nfound == 0:
				codelistList.append(codelistName)
				# print("added")
				nfound = nfound + 1
	return codelistList
	
def getCLSubsets(sheet) -> list:
	clsubsetSpecList = []
	# cmaps where the cd is a subset of a codelist will be processed as subsets.
	# some of the cd's without a subset can be processed as codelists  -- or, perhaps, the Ccode can be matched against clib info
	nfound = 0
	for r in sheet.iter_rows(min_col=1, max_col=7, values_only=True):
		ctype = r[1]
		if ctype == "Conceptual Domain":
			clsubsetSpec = r[0]
			# print("Subset label " + clsubsetSpec)
			if r[6] == 1:
				clsubsetSpec = "Subset:" + r[0]
			if clsubsetSpec not in clsubsetSpecList or nfound == 0:
				clsubsetSpecList.append(clsubsetSpec)
				# print(nfound)
				nfound = nfound + 1
	return clsubsetSpecList

# gets library authorization info from the user

	
def testRequest(reqURL:str, uname:str, passwd:str, rformat='json') -> dict:
	logging.info("API call %s", reqURL)
	r = requests.get(reqURL,headers=testhdr,auth=(uname,passwd))
	status = r.status_code
	if (status == 200):
		reqPayload = r.text
		reqJson = json.loads(reqPayload)
		return reqJson
	elif (status == 404):
		reqJson = {}
		reqJson["Status404"] = reqURL
		logging.info("%s returned %s", reqURL, status)
	else:
		reqJson = {}
		reqJson["status"] = "ErrorCode" + str(status)
		reqJson["href"] = reqURL
		errmsg = reqURL + ' returned ' + str(status) + ' (' + uname + ',' + passwd + ')'
		logging.error(errmsg)
	return reqJson	
 
def clibCodeList(varDef:dict,  tokenlist:list) -> dict:
	uname = tokenlist[0]
	passwd = tokenlist[1]
	baseURL = tokenlist[2]
	varName = varDef["name"]
	clDef = {}
	clDef["name"] = varName
	if "codelist" in varDef["_links"]:
		logging.debug("found a codelist for " + varName)
		lnx = varDef["_links"]
		codelistList = lnx["codelist"]
		if len(codelistList) > 1:
			logging.info("Multiple codelists.")
		for cl in codelistList:
			clhref = cl["href"]
			clEndPoint = baseURL + clhref
			clibRootCL = testRequest(clEndPoint, uname, passwd)
			clLnx = clibRootCL["_links"]
			clVersions = clLnx["versions"]

			""" use the most recent version of the codelist. """
			logging.info("There are " + str(len(clVersions)) + " versions for the " + varName + " codelist.")
			if len(clVersions) > 0:
				clDef["href"] = clhref
				curCLDef = clVersions[len(clVersions)-1 ]  # most recent version should be the last in returned list
				clTermsEP = baseURL + curCLDef["href"]
				clibCLDef = testRequest(clTermsEP, uname, passwd)
				clSubmissionName: str = clibCLDef["submissionValue"]
				clDef["CTVersion"] = getclparentpackage(clibCLDef)
				clDef["href"] = clhref
				clDef["name"] = clibCLDef["submissionValue"]
				clDef["terms"] = clibCLDef["terms"]
	else:
		logging.debug(varDef)
	return clDef


def lookupADaMStd(standard, version, cmap_report, tokenList, bcList) -> dict:
	""" This will find only the variables that are explcitly listed within the Library.
	Todo: Include a list of AC varnames that were not matched with a variable definition in the referenced ADaM IG.

	:param standard:
	:param version:
	:param cmap_report:
	:param tokenList:
	:param bcList:
	:return:
	"""
	uname = tokenList[0]
	passwd = tokenList[1]
	baseURL = tokenList[2]
	stdVersURL = baseURL + '/mdr/adam/' + standard + '-' + version + '/datastructures/'
	dsURL = lambda dsURL, dsName: dsURL + dsName
	varURL = lambda dsURL, varName: dsURL + '/variables/' + varName
	# this processes only 1 cmap so it creates the Define content for only 1 dataset
	# creates a list of domains because that what the existing writeXML expects
	wb = load_workbook(cmap_report)
	sheet = wb["CMAP Nodes"]
	if len(bcList) == 0:
		doConcepts = False
	else:
		doConcepts = True

	dataset = getADaMdataset(sheet)
	bcname = getBC(sheet)
	domainURL = dsURL(stdVersURL, dataset)
	domainCLIB = testRequest(domainURL, uname, passwd)
	dmdict = {}
	dmdict["OID"] = "C360.ADaM.DS." + dataset
	dmdict["Name"] = dataset
	dmdict["Label"] = domainCLIB["label"]
	dmdict["Structure"] = domainCLIB["description"]
	analysisVarsets = domainCLIB["analysisVariableSets"]

	varnames = getVarnames(sheet)
	nvarsBC = len(varnames)
	codelistRefs = getCodelists(sheet)
	clSubsets = getCLSubsets(sheet)
	clNumbers = "Number of Codelists " + str(len(codelistRefs)) + " up to " + str(len(clSubsets)) + " subsets."
	logging.info(clNumbers)
	c360varList = []
	c360clList = []
	clibOrdinal = 1
	for v in varnames:
		c360ADaMvarDef = getADaMClibVar(v, analysisVarsets)
		if v in c360ADaMvarDef :
			"""To Do: look up codelists in library json
			c360clDef = clibCodeList(c360ADaMvarDef, codelistRefs, tokenList)
			if "terms" in c360clDef:
				c360clList.append(c360clDef)
			"""
			logging.info(c360ADaMvarDef[v])
			c360ADaMvarDef["ordinal"] = str(clibOrdinal)
			c360varList.append(c360ADaMvarDef)
			clibOrdinal = clibOrdinal + 1
		else:
			logging.error("Did not find Clib info for " + v + " in " + domainURL)
	# to do process codelist, method, vlm, codelist subsets

	outcome = "For " + bcname + "(" + standard + "," + version + ")" + dataset + " has " + str(
		len(c360varList)) + " decs"
	logging.info(outcome)
	dmdict["acVarsets"] = c360varList
	aVarsetList = []
	aVarsetList.append(dmdict)

	defInfo = {}
	defInfo["ProdVersion"] = standard + "-" + version
	defInfo["Standard"] = standard
	defInfo["StdVersion"] = version
	defInfo["title"] = "CDISC 360 Analysis Concepts"
	defInfo["varsetList"] = aVarsetList
	#defInfo["codelists"] = c360clList
	#defInfo["clSubsets"] = clSubsets
	fname = dataset + "_ADaM.json"
	logging.info(fname)
	with open(fname, "w", encoding="utf-8") as fo:
		fo.write(json.dumps(defInfo))
	return (defInfo)

def getclparentpackage(clclibDef:dict) -> dict:
	if "_links" in clclibDef:
		links = clclibDef["_links"]
		return links["parentPackage"]
	return